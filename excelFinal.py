#!/usr/bin/env python3
"""
HWUM WhatsApp Blaster - Improved Version

This module implements a dual-browser WhatsApp messaging tool with a GUI.
It has been refactored for improved readability, efficiency, and standardization.
"""

import os
import sys
import re
import shutil
import tempfile
import platform
import urllib.parse
import logging
import random
import time
import threading
import queue
from pathlib import Path
from typing import Any, Dict, Optional

import pandas as pd
import psutil
from openpyxl import load_workbook

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.chrome.service import Service as ChromiumService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    WebDriverException,
)

# Webdriver managers
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.core.os_manager import ChromeType
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from webdriver_manager.chrome import ChromeDriverManager as BraveChromeDriverManager  # for Brave

# ---------------------------
# CONFIGURATION
# ---------------------------
CONFIG: Dict[str, Any] = {
    "SHEETS": {
        "LIST": "LIST",
        "MSGS": "MSGS",
        "DOCS": "DOCS",
        "MEDIA": "MEDIA",
        "SETTINGS": "SETTINGS",
    },
    "COLUMNS": {
        "LIST": {
            "sender": "Sender",
            "phone": "Phone Number",
            "name": "Name",
            "course": "Course of Interest",
            "msg_code": "Message Code",
            "doc_code": "Document Code",
            "media_code": "Media Code",
            "status": "Status",
        },
        "MSGS": {
            "msg_code": "Message Code",
            "message": "Message Encoded",
        },
        "DOCS": {
            "code": "Document Code",
            "files": ["BROCHURE_1", "BROCHURE_2", "BROCHURE_3", "BROCHURE_4"],
        },
        "MEDIA": {
            "code": "Media Code",
            "files": ["MEDIA_1", "MEDIA_2", "MEDIA_3", "MEDIA_4"],
        },
        "Settings": {
            "wd_chrome_ver": "WD_CHROME_VER",
            "wd_edge_ver": "WD_EDGE_VER",
            "wd_brave_ver": "WD_BRAVE_VER",
            "xpath_text": "XPATH_TEXT",
            "xpath_send": "XPATH_SEND",
            "xpath_attach": "XPATH_ATTACH",
            "xpath_asend": "XPATH_ASEND",
            "xpath_docs": "XPATH_DOCS",
            "xpath_media": "XPATH_MEDIA",
            "invalid_message": "INVALID_MSG",
            "min_timer": "MIN_TIMER",
            "max_timer": "MAX_TIMER",
        },
    },
    "STATUS_VALUES": {
        "INVALID": 0,
        "SENT": 1,
        "RETRY": 2,
    },
}

# ---------------------------
# Logging and Utility Functions
# ---------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("whatsapp_blaster.log"),
        logging.StreamHandler()
    ],
)

# Global queues for logging output
system_log_queue = queue.Queue()
browser1_log_queue = queue.Queue()
browser2_log_queue = queue.Queue()

def get_persistent_temp_path(instance_id: Optional[str] = None) -> str:
    """Get a persistent temporary path for storing user data."""
    base = Path(tempfile.gettempdir())
    folder = f"whatsapp_blaster_data_{instance_id}" if instance_id else "whatsapp_blaster_data"
    path = base / folder
    path.mkdir(parents=True, exist_ok=True)
    return str(path)

def async_log_worker(log_widget: tk.Text, log_q: queue.Queue) -> None:
    """Update the given text widget with messages from the log queue."""
    while True:
        try:
            message = log_q.get(timeout=1)
            log_widget.after(0, lambda: log_widget.insert(tk.END, message + "\n"))
            log_widget.after(0, log_widget.see, tk.END)
        except queue.Empty:
            continue

def log_system(message: str) -> None:
    system_log_queue.put(message)
    logging.info(message)

def log_browser(instance: int, message: str) -> None:
    if instance == 1:
        browser1_log_queue.put(message)
    elif instance == 2:
        browser2_log_queue.put(message)
    else:
        system_log_queue.put(message)

def normalize_value(value: Any) -> str:
    """Normalize a value to string; convert floats that are whole numbers to int."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()

excel_lock = threading.Lock()

def is_file_locked(filepath: str) -> bool:
    """Check if a file is currently locked."""
    try:
        with open(filepath, 'a'):
            return False
    except IOError:
        return True

def update_excel_status(excel_file: str, phone_number: str, new_status: int) -> None:
    """Update the status of a contact in the Excel file."""
    attempts = 3
    while attempts > 0:
        if is_file_locked(excel_file):
            logging.warning(f"Excel file is locked. Retrying in 2 seconds for phone {phone_number}...")
            time.sleep(2)
            attempts -= 1
        else:
            break
    with excel_lock:
        try:
            wb = load_workbook(excel_file)
            sheet = wb[CONFIG["SHEETS"]["LIST"]]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            phone_idx = headers.index(CONFIG["COLUMNS"]["LIST"]["phone"])
            status_idx = headers.index(CONFIG["COLUMNS"]["LIST"]["status"])
            updated = False
            for row in sheet.iter_rows(min_row=2):
                if normalize_value(row[phone_idx].value) == phone_number:
                    row[status_idx].value = new_status
                    updated = True
                    break
            if updated:
                wb.save(excel_file)
                logging.info(f"Updated status for {phone_number} to {new_status}")
            else:
                logging.warning(f"Phone number {phone_number} not found in Excel for updating.")
        except Exception as e:
            logging.error(f"Failed to update status for {phone_number}: {e}")

# ---------------------------
# Message Personalization
# ---------------------------
def parse_spintax(text: str) -> str:
    """
    Process spintax in the format [option1|option2|option3] and randomly select one.
    """
    pattern = re.compile(r"\[([^{}\[\]]*)\]")
    while re.search(pattern, text):
        text = re.sub(pattern, lambda m: random.choice(m.group(1).split("|")).strip(), text)
    return text

def personalize_message(encoded_template: str, contact: Dict[str, Any]) -> str:
    """
    Decode the template, substitute placeholders, process spintax, and re-encode for URL.
    """
    decoded_template = urllib.parse.unquote_plus(encoded_template)
    personalized = decoded_template.format(
        name=contact.get("Name", ""),
        sender=contact.get("Sender", ""),
        course=contact.get("Course of Interest", "")
    )
    spintax_processed = parse_spintax(personalized)
    return urllib.parse.quote_plus(spintax_processed)

# ---------------------------
# Excel Data Loader
# ---------------------------
class ExcelDataLoader:
    def __init__(self, excel_file: str) -> None:
        self.excel_file = excel_file
        self.sheets = self.load_all_sheets()

    def load_all_sheets(self) -> Dict[str, pd.DataFrame]:
        try:
            wb = load_workbook(self.excel_file)
            sheets: Dict[str, pd.DataFrame] = {}
            for sheet in wb.sheetnames:
                if sheet == CONFIG["SHEETS"]["LIST"]:
                    converters = {
                        CONFIG["COLUMNS"]["LIST"]["phone"]: str,
                        CONFIG["COLUMNS"]["LIST"]["msg_code"]: str,
                        CONFIG["COLUMNS"]["LIST"]["doc_code"]: str,
                        CONFIG["COLUMNS"]["LIST"]["media_code"]: str,
                    }
                    sheets[sheet] = pd.read_excel(self.excel_file, sheet_name=sheet, converters=converters)
                elif sheet == CONFIG["SHEETS"]["MSGS"]:
                    converters = {CONFIG["COLUMNS"]["MSGS"]["msg_code"]: str}
                    sheets[sheet] = pd.read_excel(self.excel_file, sheet_name=sheet, converters=converters)
                else:
                    sheets[sheet] = pd.read_excel(self.excel_file, sheet_name=sheet)
            return sheets
        except Exception as e:
            logging.error(f"Error loading Excel file: {e}")
            return {}

    def get_contacts(self) -> pd.DataFrame:
        df = self.sheets.get(CONFIG["SHEETS"]["LIST"])
        if df is not None:
            for col in [CONFIG["COLUMNS"]["LIST"]["phone"],
                        CONFIG["COLUMNS"]["LIST"]["msg_code"],
                        CONFIG["COLUMNS"]["LIST"]["doc_code"],
                        CONFIG["COLUMNS"]["LIST"]["media_code"]]:
                df[col] = df[col].apply(normalize_value)
        return df

    def get_messages(self) -> pd.DataFrame:
        df = self.sheets.get(CONFIG["SHEETS"]["MSGS"])
        if df is not None:
            df[CONFIG["COLUMNS"]["MSGS"]["msg_code"]] = df[CONFIG["COLUMNS"]["MSGS"]["msg_code"]].apply(normalize_value)
        return df

    def get_mapping(self, sheet_key: str, col_config: Dict[str, Any]) -> Dict[str, Any]:
        mapping: Dict[str, Any] = {}
        df = self.sheets.get(CONFIG["SHEETS"][sheet_key])
        if df is not None:
            df[col_config["code"]] = df[col_config["code"]].apply(normalize_value)
            for _, row in df.iterrows():
                key = row[col_config["code"]]
                mapping[key] = [row[col] for col in col_config["files"]]
        return mapping

    def get_docs(self) -> Dict[str, Any]:
        return self.get_mapping("DOCS", CONFIG["COLUMNS"]["DOCS"])

    def get_media(self) -> Dict[str, Any]:
        return self.get_mapping("MEDIA", CONFIG["COLUMNS"]["MEDIA"])

    def get_settings(self) -> Dict[str, Any]:
        if CONFIG["SHEETS"]["SETTINGS"] in self.sheets:
            df = self.sheets[CONFIG["SHEETS"]["SETTINGS"]]
            return dict(zip(df["Setting Name"], df["Value"]))
        return {}

# ---------------------------
# Browser Manager
# ---------------------------
class BrowserManager:
    def __init__(self, instance_id: Optional[int] = None) -> None:
        self.driver: Optional[webdriver.Chrome] = None
        self.instance_id = instance_id
        self.user_data_path = get_persistent_temp_path(str(instance_id))

    def setup_browser(self, headless: bool = False, settings: Dict[str, Any] = {}) -> webdriver.Chrome:
        if self.driver is not None:
            try:
                _ = self.driver.current_url
                return self.driver
            except Exception:
                self.driver = None

        options = Options()
        options.add_argument(f"--user-data-dir={self.user_data_path}")
        debug_port = 9222 if self.instance_id is None else 9223 + int(self.instance_id)
        options.add_argument(f"--remote-debugging-port={debug_port}")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-background-networking")
        options.add_argument("--disable-default-apps")
        options.add_argument("--disable-popup-blocking")
        options.add_argument("--no-first-run")
        options.add_argument("--no-service-autorun")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        )
        if headless:
            options.add_argument("--headless")

        browser_path, browser_type = self.locate_browser()
        if not browser_path:
            raise RuntimeError("No supported browser found on the system.")
        options.binary_location = browser_path

        # Use settings for driver version if provided
        if browser_type == "chrome":
            service = ChromeService(
                ChromeDriverManager(driver_version=settings.get(CONFIG["COLUMNS"]["Settings"]["wd_chrome_ver"])).install()
            )
        elif browser_type == "edge":
            service = EdgeService(
                EdgeChromiumDriverManager(driver_version=settings.get(CONFIG["COLUMNS"]["Settings"]["wd_edge_ver"])).install()
            )
        elif browser_type == "brave":
            service = ChromeService(
                BraveChromeDriverManager(
                    chrome_type=ChromeType.BRAVE,
                    driver_version=settings.get(CONFIG["COLUMNS"]["Settings"]["wd_brave_ver"])
                ).install()
            )
        else:
            service = ChromiumService(
                ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install()
            )
        self.driver = webdriver.Chrome(service=service, options=options)
        log_system(f"Browser instance {self.instance_id} initialized")
        return self.driver

    def locate_browser(self) -> tuple[Optional[str], Optional[str]]:
        paths = {
            "brave": [
                os.path.expandvars(r"%ProgramFiles%\BraveSoftware\Brave-Browser\Application\brave.exe"),
                os.path.expandvars(r"%ProgramFiles(x86)%\BraveSoftware\Brave-Browser\Application\brave.exe"),
                os.path.expandvars(r"%LocalAppData%\BraveSoftware\Brave-Browser\Application\brave.exe"),
                "/Applications/Brave Browser.app/Contents/MacOS/Brave Browser",
                "/usr/bin/brave-browser"
            ],
            "chrome": [
                os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
                os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
                os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe"),
                "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
                "/usr/bin/google-chrome",
                "/usr/bin/chromium",
                "/usr/bin/chromium-browser",
            ],
            "edge": [
                os.path.expandvars(r"%ProgramFiles%\Microsoft\Edge\Application\msedge.exe"),
                os.path.expandvars(r"%ProgramFiles(x86)%\Microsoft\Edge\Application\msedge.exe"),
                "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
                "/usr/bin/microsoft-edge",
            ],
        }

        for browser, paths_list in paths.items():
            for path in paths_list:
                if os.path.exists(path):
                    return path, browser
        return None, None

    def quit(self) -> None:
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None

# ---------------------------
# WhatsApp Messaging Functions
# ---------------------------
def wait_for_element(driver: webdriver.Chrome, xpath: str, timeout: int = 10) -> bool:
    """Wait for an element specified by xpath to appear on the page."""
    try:
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return True
    except TimeoutException:
        return False

def send_text_message(driver: webdriver.Chrome, phone_number: str, encoded_message: str,
                      contact: Dict[str, Any], instance_id: int, settings: Dict[str, Any]) -> Any:
    """
    Send a personalized text message to a WhatsApp contact.
    Returns True if sent, "INVALID" if number is invalid, or False otherwise.
    """
    final_message = personalize_message(encoded_message, contact)
    url = f"https://web.whatsapp.com/send?phone={phone_number}&text={final_message}"
    driver.get(url)
    
    try:
        WebDriverWait(driver, 10).until(
            lambda d: (settings[CONFIG["COLUMNS"]["Settings"]["invalid_message"]] in d.page_source) or
                      (len(d.find_elements(By.XPATH, settings[CONFIG["COLUMNS"]["Settings"]["xpath_text"]])) > 0)
        )
    except TimeoutException:
        log_system(f"[Instance {instance_id}] Timeout waiting for elements for {phone_number}.")
        return False

    if settings[CONFIG["COLUMNS"]["Settings"]["invalid_message"]] in driver.page_source:
        log_system(f"[Instance {instance_id}] Invalid phone number {phone_number}.")
        return "INVALID"

    delay = random.uniform(float(settings[CONFIG["COLUMNS"]["Settings"]["min_timer"]]),
                             float(settings[CONFIG["COLUMNS"]["Settings"]["max_timer"]]))
    time.sleep(delay)

    try:
        text_elements = driver.find_elements(By.XPATH, settings[CONFIG["COLUMNS"]["Settings"]["xpath_text"]])
        if not text_elements:
            log_system(f"[Instance {instance_id}] Message input not found for {phone_number}.")
            return False
        text_elements[0].click()
        time.sleep(0.5)
        send_buttons = driver.find_elements(By.XPATH, settings[CONFIG["COLUMNS"]["Settings"]["xpath_send"]])
        if not send_buttons:
            log_system(f"[Instance {instance_id}] Send button not found for {phone_number}.")
            return False
        send_buttons[0].click()
        log_system(f"[Instance {instance_id}] Text message sent to {phone_number}.")
        return True
    except (NoSuchElementException, Exception) as e:
        log_system(f"[Instance {instance_id}] Error sending to {phone_number}: {e}")
        return False

def attach_files(driver: webdriver.Chrome, phone_number: str, file_paths: list, xpath: str,
                 instance_id: int, settings: Dict[str, Any]) -> bool:
    """
    Attach files to a WhatsApp message.
    """
    whatsapp_url = f"https://web.whatsapp.com/send?phone={phone_number}"
    driver.get(whatsapp_url)
    try:
        if not wait_for_element(driver, settings[CONFIG["COLUMNS"]["Settings"]["xpath_attach"]], timeout=10):
            log_system(f"[Instance {instance_id}] Attach button not found for {phone_number}.")
            return False
        attach_btn = driver.find_element(By.XPATH, settings[CONFIG["COLUMNS"]["Settings"]["xpath_attach"]])
        time.sleep(1)
        attach_btn.click()
        try:
            option = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            option.click()
        except TimeoutException:
            log_system(f"[Instance {instance_id}] Option for files not clickable for {phone_number}.")
            return False
        try:
            file_input = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//input[@accept='*']"))
            )
            valid_paths = [str(p).strip() for p in file_paths if pd.notna(p)]
            if valid_paths:
                file_input.send_keys("\n".join(valid_paths))
                time.sleep(2)
                send_btn = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, settings[CONFIG["COLUMNS"]["Settings"]["xpath_asend"]]))
                )
                send_btn.click()
                log_system(f"[Instance {instance_id}] Files sent to {phone_number}.")
                return True
            else:
                log_system(f"[Instance {instance_id}] No valid files for {phone_number}.")
                return False
        except TimeoutException:
            log_system(f"[Instance {instance_id}] File input field not found for {phone_number}.")
            return False
    except Exception as e:
        log_system(f"[Instance {instance_id}] Error attaching files for {phone_number}: {e}")
        return False

def process_contact(driver: webdriver.Chrome, contact: pd.Series, messages_df: pd.DataFrame,
                    docs_mapping: Dict[str, Any], media_mapping: Dict[str, Any],
                    excel_file: str, instance_id: int, settings: Dict[str, Any]) -> None:
    """
    Process an individual contact: send message and attach files as needed.
    """
    phone = contact.get(CONFIG["COLUMNS"]["LIST"]["phone"])
    msg_code = str(contact.get(CONFIG["COLUMNS"]["LIST"]["msg_code"], "")).strip()
    doc_code = str(contact.get(CONFIG["COLUMNS"]["LIST"]["doc_code"], "")).strip()
    media_code = str(contact.get(CONFIG["COLUMNS"]["LIST"]["media_code"], "")).strip()

    if all(code == "0" for code in [msg_code, doc_code, media_code]):
        log_browser(instance_id, f"All codes zero for {phone} - Marking as SENT")
        update_excel_status(excel_file, phone, CONFIG["STATUS_VALUES"]["SENT"])
        return

    sent_successfully = False

    if msg_code != "0":
        if not messages_df.empty and msg_code in messages_df[CONFIG["COLUMNS"]["MSGS"]["msg_code"]].values:
            encoded_msg = messages_df.loc[
                messages_df[CONFIG["COLUMNS"]["MSGS"]["msg_code"]] == msg_code,
                CONFIG["COLUMNS"]["MSGS"]["message"]
            ].iloc[0]
            result = send_text_message(driver, phone, encoded_msg, contact, instance_id, settings)
            if result == "INVALID":
                update_excel_status(excel_file, phone, CONFIG["STATUS_VALUES"]["INVALID"])
                return
            sent_successfully = result
        else:
            log_browser(instance_id, f"Invalid message code {msg_code} for {phone}")

    if doc_code != "0" and doc_code in docs_mapping:
        result = attach_files(driver, phone, docs_mapping[doc_code],
                              settings[CONFIG["COLUMNS"]["Settings"]["xpath_docs"]],
                              instance_id, settings)
        sent_successfully = sent_successfully or result

    if media_code != "0" and media_code in media_mapping:
        result = attach_files(driver, phone, media_mapping[media_code],
                              settings[CONFIG["COLUMNS"]["Settings"]["xpath_media"]],
                              instance_id, settings)
        sent_successfully = sent_successfully or result

    final_status = CONFIG["STATUS_VALUES"]["SENT"] if sent_successfully else CONFIG["STATUS_VALUES"]["RETRY"]
    update_excel_status(excel_file, phone, final_status)
    time.sleep(random.uniform(
        float(settings[CONFIG["COLUMNS"]["Settings"]["min_timer"]]),
        float(settings[CONFIG["COLUMNS"]["Settings"]["max_timer"]])
    ))

# ---------------------------
# Dual-Browser WhatsApp Blaster
# ---------------------------
class DualBrowserWhatsAppBlaster:
    def __init__(self) -> None:
        self.browser_manager_1 = BrowserManager(instance_id=1)
        self.browser_manager_2 = BrowserManager(instance_id=2)
        self.log_updater_active = False
        self.thread_log_updater: Optional[threading.Thread] = None

    def start_log_updater(self, log_widget: tk.Text) -> None:
        self.log_updater_active = True

        def update_logs() -> None:
            while self.log_updater_active:
                try:
                    message = system_log_queue.get(timeout=0.1)
                    log_widget.after(0, lambda: log_widget.insert(tk.END, message + "\n"))
                    log_widget.after(0, log_widget.see, tk.END)
                except queue.Empty:
                    pass
                except Exception as e:
                    logging.error(f"Error in log updater: {e}")

        self.thread_log_updater = threading.Thread(target=update_logs, daemon=True)
        self.thread_log_updater.start()

    def stop_log_updater(self) -> None:
        self.log_updater_active = False
        if self.thread_log_updater:
            self.thread_log_updater.join(timeout=1.0)
            self.thread_log_updater = None

    def process_contacts_thread(self, driver: webdriver.Chrome, contacts_df: pd.DataFrame,
                                  messages_df: pd.DataFrame, docs_mapping: Dict[str, Any],
                                  media_mapping: Dict[str, Any], excel_file: str,
                                  instance_id: int, settings: Dict[str, Any]) -> None:
        try:
            driver.get("https://web.whatsapp.com")
            log_browser(instance_id, f"[Instance {instance_id}] Waiting for WhatsApp Web to load...")
            WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "//div[@id='pane-side']")))
            log_browser(instance_id, f"[Instance {instance_id}] WhatsApp Web loaded successfully.")
            for _, contact in contacts_df.iterrows():
                if stop_event.is_set():
                    log_browser(instance_id, f"[Instance {instance_id}] Stopping processing as requested.")
                    break
                if contact.get(CONFIG["COLUMNS"]["LIST"]["status"]) in [
                    CONFIG["STATUS_VALUES"]["INVALID"], CONFIG["STATUS_VALUES"]["SENT"]
                ]:
                    continue
                try:
                    process_contact(driver, contact, messages_df, docs_mapping, media_mapping, excel_file, instance_id, settings)
                except Exception as e:
                    log_browser(instance_id, f"[Instance {instance_id}] Error processing {contact.get(CONFIG['COLUMNS']['LIST']['phone'])}: {e}")
                    update_excel_status(excel_file, contact.get(CONFIG["COLUMNS"]["LIST"]["phone"]), CONFIG["STATUS_VALUES"]["RETRY"])
            log_browser(instance_id, f"[Instance {instance_id}] Initial processing complete.")
        except Exception as e:
            log_browser(instance_id, f"[Instance {instance_id}] Thread error: {e}")

    def retry_failed_contacts_thread(self, driver: webdriver.Chrome, retry_df: pd.DataFrame,
                                     messages_df: pd.DataFrame, docs_mapping: Dict[str, Any],
                                     media_mapping: Dict[str, Any], excel_file: str,
                                     instance_id: int, settings: Dict[str, Any]) -> None:
        try:
            if retry_df.empty:
                log_browser(instance_id, f"[Instance {instance_id}] No contacts to retry.")
                return
            driver.get("https://web.whatsapp.com")
            WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, "//div[@id='pane-side']")))
            for _, contact in retry_df.iterrows():
                if stop_event.is_set():
                    log_browser(instance_id, f"[Instance {instance_id}] Stopping retry process as requested.")
                    break
                try:
                    process_contact(driver, contact, messages_df, docs_mapping, media_mapping, excel_file, instance_id, settings)
                except Exception as e:
                    log_browser(instance_id, f"[Instance {instance_id}] Retry error for {contact.get(CONFIG['COLUMNS']['LIST']['phone'])}: {e}")
                    update_excel_status(excel_file, contact.get(CONFIG["COLUMNS"]["LIST"]["phone"]), CONFIG["STATUS_VALUES"]["INVALID"])
            log_browser(instance_id, f"[Instance {instance_id}] Retry process completed.")
        except Exception as e:
            log_browser(instance_id, f"[Instance {instance_id}] Thread error in retry: {e}")

    def send_messages(self, log_widget: tk.Text, excel_file: str, headless: bool) -> None:
        stop_event.clear()
        self.start_log_updater(log_widget)
        log_system("Starting dual browser WhatsApp blaster...")
        loader = ExcelDataLoader(excel_file)
        settings = loader.get_settings()
        if settings:
            CONFIG["WEBDRIVER_VER"] = {
                "CHROME": settings.get(CONFIG["COLUMNS"]["Settings"]["wd_chrome_ver"]),
                "EDGE": settings.get(CONFIG["COLUMNS"]["Settings"]["wd_edge_ver"]),
                "BRAVE": settings.get(CONFIG["COLUMNS"]["Settings"]["wd_brave_ver"])
            }
            CONFIG["INVALID_MSG"] = str(settings.get(CONFIG["COLUMNS"]["Settings"]["invalid_message"]))
            CONFIG["TIMER"] = {
                "MIN": float(settings.get(CONFIG["COLUMNS"]["Settings"]["min_timer"])),
                "MAX": float(settings.get(CONFIG["COLUMNS"]["Settings"]["max_timer"]))
            }
        contacts = loader.get_contacts()
        messages = loader.get_messages()
        docs = loader.get_docs()
        media = loader.get_media()
        even_contacts = contacts.iloc[::2].copy()
        odd_contacts = contacts.iloc[1::2].copy()
        log_system(f"Total contacts: {len(contacts)}, Even: {len(even_contacts)}, Odd: {len(odd_contacts)}")
        try:
            driver1 = self.browser_manager_1.setup_browser(headless, settings)
            driver2 = self.browser_manager_2.setup_browser(headless, settings)
            thread1 = threading.Thread(
                target=self.process_contacts_thread,
                args=(driver1, even_contacts, messages, docs, media, excel_file, 1, settings)
            )
            thread2 = threading.Thread(
                target=self.process_contacts_thread,
                args=(driver2, odd_contacts, messages, docs, media, excel_file, 2, settings)
            )
            thread1.start()
            thread2.start()
            thread1.join()
            thread2.join()
            log_system("Initial processing complete in both browsers.")
            loader = ExcelDataLoader(excel_file)
            contacts = loader.get_contacts()
            retry_df = contacts[contacts[CONFIG["COLUMNS"]["LIST"]["status"]] == CONFIG["STATUS_VALUES"]["RETRY"]]
            log_system(f"{len(retry_df)} contacts marked for retry.")
            if not retry_df.empty:
                even_retry = retry_df.iloc[::2].copy()
                odd_retry = retry_df.iloc[1::2].copy()
                retry_thread1 = threading.Thread(
                    target=self.retry_failed_contacts_thread,
                    args=(driver1, even_retry, messages, docs, media, excel_file, 1, settings)
                )
                retry_thread2 = threading.Thread(
                    target=self.retry_failed_contacts_thread,
                    args=(driver2, odd_retry, messages, docs, media, excel_file, 2, settings)
                )
                retry_thread1.start()
                retry_thread2.start()
                retry_thread1.join()
                retry_thread2.join()
            log_system("All processing completed. Closing browsers.")
        except Exception as e:
            log_system(f"Error in dual browser operation: {e}")
        finally:
            self.browser_manager_1.quit()
            self.browser_manager_2.quit()
            self.stop_log_updater()

# Global event to signal stop
stop_event = threading.Event()

# ---------------------------
# GUI and Main Execution
# ---------------------------
def create_gui() -> None:
    root = tk.Tk()
    root.title("HWUM WhatsApp Blaster")
    root.geometry("480x530")
    root.configure(bg="#333333")
    root.resizable(False, False)

    tk.Label(root, text="HWUM WhatsApp Blaster", bg="#333333", fg="white", font=("Arial", 20)).pack(pady=5)

    button_frame = tk.Frame(root, bg="#333333")
    button_frame.pack(pady=5)

    excel_file = tk.StringVar()
    headless_mode = tk.BooleanVar(value=False)
    blaster = DualBrowserWhatsAppBlaster()

    def download_template() -> None:
        try:
            base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
            template_path = Path(base_path) / "Template.xlsx"
            destination = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile="Template.xlsx",
                title="Save Template As"
            )
            if destination:
                shutil.copyfile(str(template_path), destination)
                messagebox.showinfo("Download Template", f"Template saved to {destination}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to download template: {e}")

    def import_excel() -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsm *.xlsx")])
        if path:
            excel_file.set(path)
            log_system(f"Excel file loaded: {path}")
            loader = ExcelDataLoader(path)
            if loader.get_settings():
                log_system("Settings successfully loaded from Excel.")
            else:
                log_system("Warning: No settings found in the Excel file.")

    def first_time_setup_wrapper() -> None:
        if not excel_file.get():
            messagebox.showerror("Error", "Please load an Excel file first.")
            return
        loader = ExcelDataLoader(excel_file.get())
        if not loader.get_settings():
            messagebox.showerror("Error", "Settings not loaded. Please check your Excel file.")
            return

        def setup_browsers() -> None:
            try:
                blaster.browser_manager_1.setup_browser(headless_mode.get(), loader.get_settings())
                blaster.browser_manager_2.setup_browser(headless_mode.get(), loader.get_settings())
                log_system("Both browser instances launched. Scan QR codes to login.")
            except Exception as e:
                log_system(f"Error initializing browsers: {e}")
        threading.Thread(target=setup_browsers, daemon=True).start()

    def send_messages_wrapper() -> None:
        if not excel_file.get():
            messagebox.showerror("Error", "Please load an Excel file first.")
            return
        threading.Thread(target=blaster.send_messages, args=(system_log_text, excel_file.get(), headless_mode.get()), daemon=True).start()

    def stop_process() -> None:
        stop_event.set()
        log_system("Stopping process...")
        blaster.browser_manager_1.quit()
        blaster.browser_manager_2.quit()
        for proc in psutil.process_iter(attrs=["pid", "name"]):
            if "chromedriver" in proc.info["name"].lower():
                try:
                    proc.terminate()
                except psutil.NoSuchProcess:
                    pass
        log_system("All processes stopped.")

    def toggle_headless() -> None:
        headless_mode.set(not headless_mode.get())
        headless_button.config(text="Headless" if headless_mode.get() else "Not Headless")

    def open_encoder() -> None:
        enc_win = tk.Toplevel()
        enc_win.title("Message Encoder")
        enc_win.resizable(False, False)
        ttk.Label(enc_win, text="Input Message:").pack(padx=10, pady=(10,0), anchor="w")
        input_text = tk.Text(enc_win, height=10, width=50)
        input_text.pack(padx=10, pady=(0,10))
        ttk.Label(enc_win, text="Encoded Message:").pack(padx=10, pady=(10,0), anchor="w")
        output_text = tk.Text(enc_win, height=10, width=50)
        output_text.pack(padx=10, pady=(0,10))
        def encode_message() -> None:
            msg = input_text.get("1.0", tk.END).strip()
            encoded = urllib.parse.quote_plus(msg)
            output_text.delete("1.0", tk.END)
            output_text.insert(tk.END, encoded)
        ttk.Button(enc_win, text="Encode Message", command=encode_message).pack(pady=(0,10))

    def delete_temp_folders() -> None:
        temp_dir = Path(tempfile.gettempdir())
        for folder in ["whatsapp_blaster_data_1", "whatsapp_blaster_data_2"]:
            folder_path = temp_dir / folder
            if folder_path.exists():
                try:
                    shutil.rmtree(folder_path)
                    messagebox.showinfo("Success", f"Deleted: {folder}")
                except Exception as e:
                    messagebox.showerror("Error", f"Could not delete {folder}: {e}")
            else:
                messagebox.showwarning("Not Found", f"{folder} does not exist.")

    ttk.Button(root, text="Encoder", command=open_encoder, width=65).pack(pady=2)
    ttk.Button(root, text="Reset", command=delete_temp_folders, width=65).pack(pady=2)
    ttk.Button(button_frame, text="Download Template", width=20, command=download_template).grid(row=0, column=0, padx=5, pady=5)
    ttk.Button(button_frame, text="Import Excel", width=20, command=import_excel).grid(row=0, column=1, padx=5, pady=5)
    ttk.Button(button_frame, text="Launch WA Web", width=20, command=first_time_setup_wrapper).grid(row=0, column=2, padx=5, pady=5)
    ttk.Button(button_frame, text="RUN", width=20, command=send_messages_wrapper).grid(row=1, column=0, padx=5, pady=5)
    ttk.Button(button_frame, text="STOP", width=20, command=stop_process).grid(row=1, column=1, padx=5, pady=5)
    headless_button = ttk.Button(button_frame, text="Not Headless", command=toggle_headless, width=20)
    headless_button.grid(row=1, column=2, padx=5, pady=5)

    tk.Label(root, text="Logs", bg="#333333", fg="white", font=("Arial", 12)).pack(anchor="w", padx=10)
    system_log_text = tk.Text(root, width=55, height=5, bg="lightyellow")
    system_log_text.pack(pady=5)
    browser1_log_text = tk.Text(root, width=55, height=5, bg="lightblue")
    browser1_log_text.pack(pady=5)
    browser2_log_text = tk.Text(root, width=55, height=5, bg="lightgreen")
    browser2_log_text.pack(pady=5)

    threading.Thread(target=async_log_worker, args=(browser1_log_text, browser1_log_queue), daemon=True).start()
    threading.Thread(target=async_log_worker, args=(browser2_log_text, browser2_log_queue), daemon=True).start()

    tk.Label(root, text="qt3000@hw.ac.uk", bg="#333333", fg="white", font=("Arial", 10)).pack(side=tk.BOTTOM, pady=1)

    def cleanup() -> None:
        blaster.browser_manager_1.quit()
        blaster.browser_manager_2.quit()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", cleanup)
    root.mainloop()

if __name__ == "__main__":
    create_gui()
